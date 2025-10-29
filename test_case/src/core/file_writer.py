import threading
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from src.util.logging_util import get_logger

logger = get_logger(__name__)

# 全局锁
write_lock = threading.Lock()

class ExcelWriter:
    """Excel文件写入器"""
    
    def __init__(self, settings):
        self.settings = settings
        # 样式配置
        self.font_name = settings.get_config_value("output_excel_style.font_name")
        self.font_size = settings.get_config_value("output_excel_style.font_size")
        self.first_column_width = settings.get_config_value("output_excel_style.first_column_width")
        self.other_columns_width = settings.get_config_value("output_excel_style.other_columns_width")
        
        # 对齐样式配置
        self.header_row_style = settings.get_config_value("output_excel_style.header_row_style")
        self.first_column_style = settings.get_config_value("output_excel_style.first_column_style")
        self.other_columns_style = settings.get_config_value("output_excel_style.other_columns_style")
        
        # 获取数据起始行
        self.data_start_row = settings.get_config_value("input_excel_processing.data_start_row")
    
    def _remap_output_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """直接对原始输出进行映射，每行一条测试用例"""
        if not data:
            return []
        
        remapped_data = []
        
        for index, item in enumerate(data, 1):
            # 计算原始行号，从data_start_row开始
            original_row_num = item.get("原始行号", 0)
            if original_row_num > 0:
                display_row_num = original_row_num + self.data_start_row - 1
            else:
                display_row_num = ""
                
            remapped_item = {
                "序号": index,
                "原始行号": display_row_num,
                "需求名称": item.get("需求名称", ""),
                "测试点编号": item.get("测试点编号", ""),
                "测试点": item.get("测试点", ""),
                "前置条件": item.get("前置条件", ""),
                "测试步骤": item.get("测试步骤", ""),
                "预期结果": item.get("预期结果", "")
            }
            remapped_data.append(remapped_item)
        
        return remapped_data
    
    def write_data(self, data_dict: Dict[str, List[Dict[str, Any]]], output_path: Path) -> bool:
        """写入Excel文件"""
        try:
            if not data_dict:
                logger.warning("没有数据可写入")
                return False
            
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            logger.debug(f"确保输出目录存在: {output_path.parent}")
            
            # 重映射数据
            remapped_data_dict = {}
            for sheet_name, data in data_dict.items():
                remapped_data = self._remap_output_data(data)
                remapped_data_dict[sheet_name] = remapped_data
            
            # 使用全局锁写入Excel文件
            with write_lock:
                self._write_formatted_excel(remapped_data_dict, str(output_path))
            
            return True
            
        except Exception as e:
            logger.error(f"写入Excel文件失败: {e}")
            return False
    
    def _write_formatted_excel(self, remapped_data_dict: Dict[str, List[Dict[str, Any]]], output_path: str):
        """写入格式化的Excel文件并应用样式"""
        try:
            logger.debug(f"开始写入Excel文件: {output_path}")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, data in remapped_data_dict.items():
                    df = pd.DataFrame(data)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # 应用Excel样式
                    self._apply_excel_styling(worksheet, df)
            
            logger.info(f"已生成格式化的Excel文件: {output_path}")
            
        except Exception as e:
            logger.error(f"写入Excel失败: {e}")
            raise
    
    def _apply_excel_styling(self, worksheet, df):
        """应用Excel样式"""
        # 创建字体
        font = Font(name=self.font_name, size=self.font_size)
        bold_font = Font(name=self.font_name, size=self.font_size, bold=True)
        
        # 设置列宽
        for col_idx, col in enumerate(worksheet.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_idx == 1:  # 第一列
                worksheet.column_dimensions[col_letter].width = self.first_column_width
            else:  # 其他列
                worksheet.column_dimensions[col_letter].width = self.other_columns_width
        
        # 创建对齐样式
        header_alignment = Alignment(
            horizontal=self.header_row_style.get("horizontal", "center"),
            vertical=self.header_row_style.get("vertical", "center"),
            wrap_text=True
        )
        
        first_column_alignment = Alignment(
            horizontal=self.first_column_style.get("horizontal", "center"),
            vertical=self.first_column_style.get("vertical", "center"),
            wrap_text=True
        )
        
        other_columns_alignment = Alignment(
            horizontal=self.other_columns_style.get("horizontal", "left"),
            vertical=self.other_columns_style.get("vertical", "center"),
            wrap_text=True
        )
        
        # 设置所有单元格的字体
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.font = font
        
        # 设置表头样式（第一行）- 水平居中，垂直居中
        for cell in worksheet[1]:
            cell.font = bold_font
            cell.alignment = header_alignment
        
        # 设置第一列样式 - 水平居中，垂直居中
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = first_column_alignment
        
        # 设置其他列样式 - 水平左对齐，垂直居中，max_col=len(df.columns)+1
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=len(df.columns)+1):
            for cell in row:
                cell.alignment = other_columns_alignment
        
        logger.debug(f"已应用Excel样式到工作表")

class FileWriterFactory:
    """文件写入器工厂"""
    
    @staticmethod
    def create_file_writer(writer_type: str = "excel", settings=None, **kwargs):
        """创建文件写入器"""
        if writer_type == "excel":
            return ExcelWriter(settings=settings, **kwargs)
        else:
            raise ValueError(f"不支持的写入器类型: {writer_type}")