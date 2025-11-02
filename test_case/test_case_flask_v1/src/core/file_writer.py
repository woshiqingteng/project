"""
文件写入模块
处理格式化的Excel输出生成
"""

import threading
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

from ..util.logger import get_logger


logger = get_logger(__name__)
_write_lock = threading.Lock()


class ExcelWriter:
    """具有格式化功能的Excel文件写入器"""
    
    def __init__(self, settings):
        """使用样式配置初始化写入器"""
        self._settings = settings
        self._font_name = settings.get("output_excel_style.font_name")
        self._font_size = settings.get("output_excel_style.font_size")
        self._first_col_width = settings.get("output_excel_style.first_column_width")
        self._other_cols_width = settings.get("output_excel_style.other_columns_width")
        self._data_start_row = settings.get("input_excel_processing.data_start_row")
    
    def write(self, data_dict: Dict[str, List[Dict[str, Any]]], output_path: Path) -> bool:
        """将数据写入格式化的Excel文件
        
        Args:
            data_dict: 映射表名到数据的字典
            output_path: 输出文件路径
            
        Returns:
            成功返回True，否则返回False
        """
        if not data_dict:
            logger.warning("没有数据可写入")
            return False
        
        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            with _write_lock:
                self._write_excel(data_dict, str(output_path))
            return True
        except Exception as e:
            logger.error(f"Excel写入失败: {e}")
            return False
    
    def _write_excel(self, data_dict: Dict[str, List[Dict[str, Any]]], output_path: str) -> None:
        """内部Excel写入实现"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, data in data_dict.items():
                df = self._prepare_dataframe(data)
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                self._apply_styling(worksheet, df)
        
        logger.info(f"已生成格式化的Excel文件: {output_path}")
    
    def _prepare_dataframe(self, data: List[Dict[str, Any]]) -> pd.DataFrame:
        """准备具有适当列映射的DataFrame"""
        prepared_data = []
        
        for idx, item in enumerate(data, 1):
            original_row = item.get("原始行号", 0)
            display_row = original_row + self._data_start_row - 1 if original_row > 0 else ""
            
            prepared_data.append({
                #"序号": idx,
                #"原始行号": display_row,
                "L4项目": item.get("测试点", ""),
                #"测试点编号": item.get("测试点编号", ""),
                "三级项目": item.get("测试点描述", ""),
                #"前置条件": item.get("前置条件", ""),
                "测试方法": item.get("测试步骤", ""),
                "预判定标准": item.get("预期结果", "")
            })
        
        return pd.DataFrame(prepared_data)
    
    def _apply_styling(self, worksheet, df: pd.DataFrame) -> None:
        """对工作表应用格式和样式"""
        # 设置列宽
        for col_idx, col in enumerate(worksheet.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            width = self._first_col_width if col_idx == 1 else self._other_cols_width
            worksheet.column_dimensions[col_letter].width = width
        
        # 创建样式
        font = Font(name=self._font_name, size=self._font_size)
        bold_font = Font(name=self._font_name, size=self._font_size, bold=True)
        
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        first_col_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        other_cols_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 应用字体
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.font = font
        
        # 表头行
        for cell in worksheet[1]:
            cell.font = bold_font
            cell.alignment = header_align
        
        # 第一列
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = first_col_align
        
        # 其他列
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=len(df.columns)+1):
            for cell in row:
                cell.alignment = other_cols_align


class FileWriterFactory:
    """文件写入器工厂"""
    
    @staticmethod
    def create(writer_type: str = "excel", settings=None, **kwargs):
        """创建文件写入器实例
        
        Args:
            writer_type: 要创建的写入器类型
            settings: 配置设置
            **kwargs: 附加参数
            
        Returns:
            文件写入器实例
            
        Raises:
            ValueError: 如果写入器类型不受支持
        """
        if writer_type == "excel":
            return ExcelWriter(settings=settings, **kwargs)
        else:
            raise ValueError(f"不支持的写入器类型: {writer_type}")